import openpyxl
from tkinter import Tk, filedialog, Button, Label, StringVar, Text, Scrollbar, RIGHT, Y
import tkinter as tk
from openpyxl.styles import PatternFill

# Global variables to hold task headings, steps, and the current index for navigation
task_headings = []
task_steps = []
current_task_index = 0

# Function to open file dialog and get dictionary from the Python file
def open_file():
    file_path = filedialog.askopenfilename(title="Select Python File", filetypes=(("Python files", "*.py"),))
    if file_path:
        try:
            globals_dict = {}
            # Ensure the Python file is read with UTF-8 encoding to avoid charmap issues
            with open(file_path, encoding='utf-8') as f:
                exec(f.read(), globals_dict)
            global action_examples
            action_examples = globals_dict['action_examples']
            status_label.config(text="File loaded successfully!", fg="green")

            # Extract task headings and steps
            extract_task_details(action_examples)

            # Write to Excel
            write_to_excel(task_headings, 'Task_Headings.xlsx')

            # Display the first task
            display_task()

        except Exception as e:
            status_label.config(text=f"Error: {str(e)}", fg="red")
    else:
        status_label.config(text="No file selected.", fg="red")


# Extract task headings and steps from the dictionary
def extract_task_details(action_dict):
    global task_headings, task_steps
    task_headings = [key for key in action_dict.keys()]
    task_steps = [value for value in action_dict.values()]

# Write task headings to Excel
def write_to_excel(task_headings, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Task Headings"

    # Write the task headings to the first column
    for idx, heading in enumerate(task_headings, start=1):
        sheet[f'A{idx}'] = heading

    workbook.save(output_file)

# Function to update Excel with 'Correct' or 'Update' with colors
def update_excel(choice):
    workbook = openpyxl.load_workbook('Task_Headings.xlsx')
    sheet = workbook.active

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Update based on user's choice (Correct or Update)
    row = current_task_index + 1  # Task index corresponds to the Excel row
    sheet[f'C{row}'] = choice
    if choice == 'Correct':
        sheet[f'C{row}'].fill = green_fill
    else:
        sheet[f'C{row}'].fill = yellow_fill

    workbook.save('Task_Headings.xlsx')
    status_label.config(text=f"{choice} has been applied to the task in row {row}.", fg="blue")

# Display the current task (heading and steps)
def display_task():
    if task_headings and task_steps:
        task_label.set(f"Task: {task_headings[current_task_index]}")
        task_steps_display.delete(1.0, tk.END)  # Clear the previous steps
        steps = task_steps[current_task_index]
        formatted_steps = format_steps(steps)
        task_steps_display.insert(tk.END, formatted_steps)

# Format the task steps for display
def format_steps(steps):
    formatted = ""
    for step in steps:
        if isinstance(step, dict):
            formatted += f"- {list(step.keys())[0]}: {list(step.values())[0]}\n"
        else:
            formatted += f"- {step}\n"
    return formatted

# Function to navigate to the next task
def next_task():
    global current_task_index
    if current_task_index < len(task_headings) - 1:
        current_task_index += 1
        display_task()

# Function to navigate to the previous task
def previous_task():
    global current_task_index
    if current_task_index > 0:
        current_task_index -= 1
        display_task()

# GUI setup
# Initialize Tkinter window
root = tk.Tk()
root.title("Task Dictionary to Excel")

# Set the window size and disable resizing
root.geometry("500x500")
root.configure(bg="#f0f8ff")  # Light background color
root.resizable(False, False)  # Disable window resizing

# Load file button
load_button = tk.Button(root, text="📂 Load Python File", command=open_file, font=("Helvetica", 12, "bold"), bg="#4682b4", fg="white", padx=10, pady=5)
load_button.pack(pady=10)

# Status label
status_label = tk.Label(root, text="", font=("Arial", 10, "italic"), bg="#f0f8ff")
status_label.pack(pady=10)

# Task heading display
task_label = StringVar()
task_display = tk.Label(root, textvariable=task_label, font=("Helvetica", 16, "bold"), fg="#2e8b57", bg="#f0f8ff", wraplength=490)
task_display.pack(pady=10)


# Task steps display (with scrollbar) with wrapping enabled
task_steps_display = Text(root, height=10, width=50, wrap=tk.WORD, font=("Arial", 12), padx=10, pady=5)
task_steps_display.pack(pady=10)

# Scrollbar for the task steps display
scrollbar = Scrollbar(root, command=task_steps_display.yview)
task_steps_display.configure(yscrollcommand=scrollbar.set)
scrollbar.pack(side=RIGHT, fill=Y)

# Navigation buttons with arrow symbols
prev_button = tk.Button(root, text="⬅ Previous", command=previous_task, font=("Helvetica", 12, "bold"), bg="#ff6347", fg="white", padx=10, pady=5)
prev_button.pack(side=tk.LEFT, padx=20, pady=10)

next_button = tk.Button(root, text="Next ➡", command=next_task, font=("Helvetica", 12, "bold"), bg="#32cd32", fg="white", padx=10, pady=5)
next_button.pack(side=tk.RIGHT, padx=20, pady=10)

# Buttons for "Correct" and "Update"
correct_button = tk.Button(root, text="✔ Correct", command=lambda: update_excel("Correct"), font=("Helvetica", 12, "bold"), bg="#228b22", fg="white", padx=10, pady=5)
correct_button.pack(pady=5)

update_button = tk.Button(root, text="✍ Update", command=lambda: update_excel("Update"), font=("Helvetica", 12, "bold"), bg="#ffa500", fg="white", padx=10, pady=5)
update_button.pack(pady=5)

# Main GUI loop
root.mainloop()
